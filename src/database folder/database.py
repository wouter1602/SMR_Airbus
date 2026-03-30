"""
database.py


Excel persistence layer for the barcode logging system.

"""


# ---------------------------------------------------
# IMPORTS
# ---------------------------------------------------

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------

MAX_BARCODES_PER_WING = 16


# ---------------------------------------------------
# LOGGER
# ---------------------------------------------------

logger = logging.getLogger("barcode_system.database")


# ---------------------------------------------------
# EXCEL MANAGER
# ---------------------------------------------------

class ExcelManager:
    """
    Handles Excel file creation and barcode entry writing.

    Wing IDs are 1-based integers. Each wing occupies exactly
    one row in the spreadsheet, starting at row 2 (row 1 = headers).
    """

    WING_ID_COLUMN      = 1
    TIMESTAMP_COLUMN    = 2
    FIRST_BARCODE_COLUMN = 3   # Barcode_1 starts here

    def __init__(self, filename: str):
        self.filename = filename
        self._ensure_file_exists()

    def _get_or_create_row(self, ws, wing_id: str) -> int:
        """Find existing row for wing_id or append a new one."""
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == wing_id:
                return row[0].row
        # Not found — use next empty row
        return ws.max_row + 1
    # ---------------------------------------------------

    def _ensure_file_exists(self):
        """Create the Excel workbook with a header row if it doesn't exist."""
        if os.path.exists(self.filename):
            logger.info(f"Excel file found: {self.filename}")
            return

        logger.info(f"Creating new Excel file: {self.filename}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Barcodes"

        headers = ["Wing_ID", "Timestamp"]
        for i in range(1, MAX_BARCODES_PER_WING + 1):
            headers.append(f"Barcode_{i}")
            headers.append(f"Inspection_{i}")

        ws.append(headers)
        wb.save(self.filename)

    # ---------------------------------------------------

    def _barcode_column(self, slot: int) -> int:
        """
        Return the Excel column index for a given 1-based slot number.

        Formula:
          barcode_col    = FIRST_BARCODE_COLUMN + (slot - 1) * 2
          inspection_col = barcode_col + 1

        Example for slot 1:  barcode → col 3,  inspection → col 4
        Example for slot 2:  barcode → col 5,  inspection → col 6
        """
        if not (1 <= slot <= MAX_BARCODES_PER_WING):
            raise ValueError(
                f"Slot {slot} is out of range (1–{MAX_BARCODES_PER_WING})."
            )
        return self.FIRST_BARCODE_COLUMN + (slot - 1) * 2

    # ---------------------------------------------------

    def write_barcode(
        self,
        wing_id: str,
        slot: int,
        barcode: str,
        inspection: str,
    ):
        """
        Write a barcode and its inspection result to the spreadsheet.

        Args:
            wing_id:    1-based wing identifier
            slot:       1-based slot number (1–16)
            barcode:    scanned barcode string
            inspection: "PASS" or "FAIL"
        """
        wb = load_workbook(self.filename)
        ws = wb.active

        row = self._get_or_create_row(ws, wing_id)
        barcode_col    = self._barcode_column(slot)
        inspection_col = barcode_col + 1

        ws.cell(row=row, column=self.WING_ID_COLUMN).value   = wing_id
        ws.cell(row=row, column=self.TIMESTAMP_COLUMN).value = datetime.now()
        ws.cell(row=row, column=barcode_col).value           = barcode
        ws.cell(row=row, column=inspection_col).value        = inspection

        wb.save(self.filename)

        logger.info(
            f"Excel updated | Wing {wing_id} | Slot {slot} | "
            f"Barcode {barcode} | {inspection}"
        )
   
