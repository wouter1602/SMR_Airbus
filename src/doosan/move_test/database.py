#!/usr/bin/env python3
"""
Functions for managing the wing panel database.
"""
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from typing import Tuple, Any
import asyncio

FILE_NAME = "wing_panel.xlsx"

POSITIONS = [
    "1-1", "1-2", "1-3", "1-4", "1-5", "1-6", "1-7", "1-8",
    "2-1", "2-2", "2-3", "2-4", "2-5", "2-6", "2-7", "2-8",
] # Wing positions.


class WingPanelDB:
    """
    Class for managing the wing panel database.
    """
    def __init__(self, file_name: str = FILE_NAME):
        """
        Initialize the database with the given file name.

        Params:
            file_name (str): The name of the Excel file to use for the database.
        """
        self.file_name = file_name
        self._current_row: int | None = None
        self._current_wing: str | None = None

    def _load_sheet(self) -> Tuple[Workbook, Any]:
        """
        Load the Excel sheet for the database.

        Returns:
            Tuple[Workbook, Any]: The workbook and active sheet.
        """
        if os.path.exists(self.file_name):
            workbook = load_workbook(self.file_name)
        else:
            workbook = Workbook()
            sheet = workbook.active or workbook.create_sheet()
            sheet.append(["WING ID", "Timestamp"] + POSITIONS)

        return workbook, workbook.active

    async def start_wing(self, wing_id: str):
        """
        Start a new wing by adding a row to the database. (Asynchronously)

        Params:
            wing_id (str): The ID of the wing to start.
        """
        await asyncio.to_thread(self._start_wing, wing_id)

    def _start_wing(self, wing_id: str):
        """
        Start a new wing by adding a row to the database. (Not asynchronously)

        Params:
            wing_id (str): The ID of the wing to start.
        """
        workbook, sheet = self._load_sheet()

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([wing_id, timestamp] + [""] * len(POSITIONS))

        self._current_row = sheet.max_row
        self._current_wing = wing_id

        workbook.save(self.file_name)

    async def add_panel(self, position: str, barcode: str):
        """
        Add a panel to the current wing. (Asynchronously)

        Params:
            position (str): The position of the panel.
            barcode (str): The barcode of the panel.
        """
        await asyncio.to_thread(self._add_panel, position, barcode)

    def _add_panel(self, position: str, barcode: str) -> None:
        """
        Add a panel to the current wing. (Not asynchronously)

        Params:
            position (str): The position of the panel.
            barcode (str): The barcode of the panel.
        """
        if self._current_row is None:
            raise RuntimeError("No active wing. Call start_wing() first.")

        if position not in POSITIONS:
            raise ValueError(f"Invalid position: {position!r}. Must be one of {POSITIONS}.")

        workbook, sheet = self._load_sheet()

        col_index = POSITIONS.index(position) + 3  # offset: Wing ID + Timestamp
        sheet.cell(row=self._current_row, column=col_index).value = barcode

        workbook.save(self.file_name)

    @property

    def current_wing(self) -> str | None:
        """
        Get the current wing ID.

        Returns:
            str | None: The current wing ID, or None if no wing is active.
        """
        return self._current_wing
