"""
Cognex In-Sight camera – simple Python driver
Connects via ISNM (TCP port 10000)
"""
import numpy as np
import socket
import time
import urllib.request
import io
import os
from enum import IntEnum
from openpyxl import Workbook, load_workbook
from datetime import datetime
try:
    from PIL import Image
except ImportError:
    Image = None



class CognexCamera:

    def __init__(self, ip, username="admin", password=""):
        self.ip = ip
        self.username = username
        self.password = password
        self._sock = None

    # ── connection ────────────────────────────────────────────────────

    def connect(self):
        self._sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self._sock.settimeout(5)
        self._sock.connect((self.ip, 23))

        self._read_until(b"User: ")
        self._sock.sendall((self.username + "\r\n").encode())

        after_user = self._read_until(b"Password: ", b"User Logged In", b"Invalid")
        if b"User Logged In" in after_user:
            return

        self._sock.sendall((self.password + "\r\n").encode())
        after_pass = self._read_until(b"User Logged In", b"Invalid Password")

        if b"User Logged In" not in after_pass:
            raise ConnectionError("Login failed")

        self._send("SO0")
        time.sleep(0.3)
        self._send("SO1")
        time.sleep(0.8)
        print('connected')

    def disconnect(self):
        if self._sock:
            self._sock.close()
            self._sock = None

    # ── camera actions ────────────────────────────────────────────────

    def trigger(self):
        resp = self._send("SW8")
        return resp.strip() == "1"
        time.sleep(0.3)

    def read_cell(self, cell):
        cell = cell.strip().upper()
        col = "".join(c for c in cell if c.isalpha())
        row = "".join(c for c in cell if c.isdigit()) or "0"
        command = f"GV{col}{int(row):03d}"

        self._sock.sendall((command + "\r\n").encode())
        raw = self._read_lines(2)
        lines = raw.decode(errors="replace").strip().splitlines()

        if not lines or lines[0].strip() != "1":
            return None
        return lines[1].strip() if len(lines) >= 2 else None

    def trigger_and_read_box(self, cells):
        self.trigger()
        time.sleep(1.0)
        return {cell: self.read_cell(cell) for cell in cells}

    def trigger_and_read_scan(self, cells):
        self.trigger()
        time.sleep(2.0)
        return {cell: self.read_cell(cell) for cell in cells}
    def set_online(self):
        self._send("SO1")
        time.sleep(0.5)

    def set_offline(self):
        self._send("SO0")

    def get_image(self):
        if Image is None:
            raise ImportError("Install Pillow first:  pip install Pillow")

        opener = urllib.request.build_opener()
        if self.username:
            pm = urllib.request.HTTPPasswordMgrWithDefaultRealm()
            pm.add_password(None, f"http://{self.ip}/", self.username, self.password)
            opener.add_handler(urllib.request.HTTPBasicAuthHandler(pm))

        for path in ["/image.bmp", "/snapshot", "/GetImage.cgi", "/image.jpg"]:
            try:
                with opener.open(f"http://{self.ip}{path}", timeout=4) as r:
                    return Image.open(io.BytesIO(r.read()))
            except Exception:
                continue
        raise ConnectionError(f"Could not fetch image from {self.ip}")

    def change_job_box(self):
        self._send("SIA0261")
    def change_job_scan(self):
        self._send("SIA0260")
    def show_image(self):
        self.get_image().show()

    # ── internals ─────────────────────────────────────────────────────

    def _send(self, command):
        self._sock.sendall((command + "\r\n").encode())
        self._sock.settimeout(5)
        buf = b""
        try:
            while True:
                chunk = self._sock.recv(4096)
                if not chunk:
                    break
                buf += chunk
                if buf.endswith(b"\r\n"):
                    break
        except socket.timeout:
            pass
        return buf.decode(errors="replace").strip()

    def _read_until(self, *markers, timeout=4.0):
        self._sock.settimeout(0.5)
        buf = b""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                buf += self._sock.recv(4096)
                if any(m in buf for m in markers):
                    break
            except socket.timeout:
                continue
        return buf

    def _read_lines(self, n, timeout=5.0):
        self._sock.settimeout(0.5)
        buf = b""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                buf += self._sock.recv(4096)
                if buf.decode(errors="replace").strip().count("\n") >= n - 1:
                    break
            except socket.timeout:
                continue
        return buf
    def set_job_by_id(self, job_id):
        if not isinstance(job_id, int):
            raise ValueError("job_id must be an integer")
        if job_id < 0 or job_id > 999:
            raise ValueError("job_id must be between 0 and 999")

        self.set_offline()
        time.sleep(0.5)

        # stuur command zonder _send (die kapt te vroeg)
        self._sock.sendall((f"SJ{job_id}\r\n").encode())

        # wacht expliciet op response
        response = self._read_lines(1, timeout=3).decode(errors="replace").strip()
        print(f"SJ response raw: '{response}'")

        if not response:
            raise RuntimeError("No response from camera (timeout)")

        try:
            status = int(response)
        except ValueError:
            raise RuntimeError(f"Unexpected response from camera: {response}")

        if status != 1:
            raise RuntimeError(f"SJ failed with status {status}")

        return True
class PickupType(IntEnum):
    Empty = 0
    Paper = 1
    Foam = 2
    No_detect = 3
    wrong_panel = 4
    right_panel = 5
def scan_box_3():
    def convert(value):
        if value == '#ERR':
            return '#ERR'
        elif float(value) == 1.0:
            return 1
        else:
            return 0
    # read right cells
    cam.change_job_box()
    position_cells = ["A42", "B42", "C42", "D42", "E42", "F42", "G42"]  # cel1 t/m cel5
    data_scan_box = cam.trigger_and_read_box(position_cells)

    # convert result
    paper = convert(data_scan_box['A42'])
    panel_M = convert(data_scan_box['B42'])
    found = convert(data_scan_box['C42'])
    panel = convert(data_scan_box['D42'])
    empty = convert(data_scan_box['E42'])
    foam = convert(data_scan_box['F42'])
    left_gap = convert(data_scan_box['G42'])

    #decide output
    if empty == 1:
        print("panel 3 is empty, please refill")
        # stop robot
        # change light
        return PickupType.Empty

    elif paper==1:
        return PickupType.Paper
        #do (3_bin)
    elif foam ==1:
        return PickupType.Foam
        # do (3_bin)
    elif panel == 1:
        if found == 1:
            if panel_M == 1:
                if left_gap == 1:
                    return PickupType.right_panel
                else:
                    return PickupType.wrong_panel
            else:
                return PickupType.wrong_panel
        else:
            return PickupType.No_detect

def scan_box_2():

    def convert(value):
        if value == '#ERR':
            return '#ERR'
        elif float(value) == 1.0:
            return 1
        else:
            return 0
    # read right cells
    cam.change_job_box()
    position_cells = ["A42", "B42", "C42", "D42", "E42", "F42", "G42"]  # cel1 t/m cel5
    data_scan_box = cam.trigger_and_read_box(position_cells)

    # convert result
    paper = convert(data_scan_box['A42'])
    panel_M = convert(data_scan_box['B42'])
    found = convert(data_scan_box['C42'])
    panel = convert(data_scan_box['D42'])
    empty = convert(data_scan_box['E42'])
    foam = convert(data_scan_box['F42'])
    left_gap = convert(data_scan_box['G42'])
    #decide output
    if empty == 1:
        print("panel 3 is empty, please refill")
        # stop robot
        # change light
        return PickupType.Empty

    elif paper==1:

        return PickupType.Paper
        #do (3_bin)
    elif foam ==1:
        return PickupType.Foam
        # do (3_bin)
    elif panel == 1:
        if found == 1:
            if panel_M == 1:
                if left_gap == 0:
                    return PickupType.right_panel
                else:
                    return PickupType.wrong_panel
            else:
                return PickupType.wrong_panel
        else:
            return PickupType.No_detect

def scan_box_1():
    def convert(value):
        if value == '#ERR':
            return '#ERR'
        elif float(value) == 1.0:
            return 1
        else:
            return 0
    # read right cells
    cam.change_job_box()
    position_cells = ["A42", "D42", "E42", "F42",]  # cel1 t/m cel5
    data_scan_box = cam.trigger_and_read_box(position_cells)

    # convert result
    paper = convert(data_scan_box['A42'])
    panel = convert(data_scan_box['D42'])
    empty = convert(data_scan_box['E42'])
    foam = convert(data_scan_box['F42'])

    if empty == 1:
        print("panel 1 is empty, please refill")
        # stop robot
        # change light
        return PickupType.Empty

    elif paper==1:
        return PickupType.Paper
        #do (3_bin)
    elif foam ==1:
        return PickupType.Foam
        # do (3_bin)
    elif panel == 1:
        cam.change_job_scan()
        position_cells_scan = ["Q35"]
        data = cam.trigger_and_read_scan(position_cells_scan)
        length = int(float(data['Q35']))
        print(length)
        if 325<length<370:
            position_cells_scan_1 = ["U38", "U39", "U40", "U41", "U42", "U43"]
            data = cam.trigger_and_read_scan(position_cells_scan_1)
            values = np.array([float(data[c]) if data[c] is not None else np.nan for c in position_cells_scan_1],dtype=np.float32)
            return PickupType.right_panel, values
        else:
            return PickupType.wrong_panel

    else:
        return PickupType.No_detect

def scan_3():
    cam.change_job_scan()
    position_cells = ["S38", "S39", "S40", "S41", "S42", "S43", "S44"]
    data = cam.trigger_and_read_scan(position_cells)
    values = np.array([float(data[c]) if data[c] is not None else np.nan for c in position_cells],dtype=np.float32)
    return values

def scan_2():
    cam.change_job_scan()
    position_cells = ["T38", "T39", "T40", "T41", "T42", "T43", "T44"]
    data = cam.trigger_and_read_scan(position_cells)
    values = np.array([float(data[c]) if data[c] is not None else np.nan for c in position_cells],dtype=np.float32)
    return values
#def scan_1():


def log_wing(wing_id):
    # Check if file exists
    if os.path.exists(FILE_NAME):
        workbook = load_workbook(FILE_NAME)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers
        sheet.append(["Wing ID", "Timestamp"])

    # Get current timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Append data (new row automatically)
    sheet.append([wing_id, timestamp])

    # Save file
    workbook.save(FILE_NAME)




# Fixed positions
POSITIONS = [
    "1-1","1-2","1-3","1-4","1-5","1-6","1-7","1-8",
    "2-1","2-2","2-3","2-4","2-5","2-6","2-7","2-8"
]

current_row = None
current_wing = None


def _load_sheet():
    if os.path.exists(FILE_NAME):
        wb = load_workbook(FILE_NAME)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active

        # Header
        headers = ["Wing ID", "Timestamp"] + POSITIONS
        sheet.append(headers)

    return wb, sheet
def start_wing(wing_id):
    global current_row, current_wing

    wb, sheet = _load_sheet()

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Create empty row for wing + 16 positions
    row_data = [wing_id, timestamp] + [""] * len(POSITIONS)
    sheet.append(row_data)

    current_row = sheet.max_row
    current_wing = wing_id

    wb.save(FILE_NAME)
def add_panel(position, barcode):
    global current_row

    if current_row is None:
        raise Exception("Start a wing first using start_wing(wing_id)")

    if position not in POSITIONS:
        raise ValueError(f"Invalid position: {position}")

    wb, sheet = _load_sheet()

    # Find column index (offset by 2 because of Wing ID + Timestamp)
    col_index = POSITIONS.index(position) + 3

    sheet.cell(row=current_row, column=col_index).value = barcode

    wb.save(FILE_NAME)
#------
FILE_NAME = "wing_panels.xlsx"



cam = CognexCamera("192.168.0.12", username="admin", password="")
cam2= CognexCamera("192.168.0.10", username='admin', password="")

cam.connect()
cam2.connect()
cam.trigger()
cam.get_image()
cam.show_image()
cam2.trigger()
start_wing(123)
add_panel("1-1", 123)
#print(scan_box_1())



#output=scan_box_3()
#print(output)
# move to scan position --> only if output == 5 / panel.right else stop or move to bin
#pickup=scan_3()
#print(pickup)
cam.disconnect()
cam2.disconnect()
