"""
Industrial Barcode Logging System
----------------------------------

Reads barcode data from Cognex scanner via Modbus TCP
and logs them into an Excel sheet.

System responsibilities:
  1. Wait for barcode scan
  2. Read barcode from registers
  3. Determine placement slot (1-16)
  4. Perform inspection logic
  5. Store barcode + inspection result in Excel

Robot placement is handled by another system.

Author: Example Industrial System
"""

# ---------------------------------------------------
# IMPORTS
# ---------------------------------------------------

import time
import logging
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from pymodbus.client import ModbusTcpClient
from pymodbus.exceptions import ModbusException


# ---------------------------------------------------
# SYSTEM CONFIGURATION
# ---------------------------------------------------

SCANNER_IP = "127.0.0.1" # Replace with actual scanner IP address
MODBUS_PORT = 5020 # Standard Modbus TCP port

MAX_BARCODES_PER_WING = 16

EXCEL_FILE = "wing_barcode_database.xlsx"

SCAN_FLAG_REGISTER = 2
BARCODE_START_REGISTER = 10
BARCODE_REGISTER_LENGTH = 20
SCANNER_RESET_REGISTER = 3

SCAN_POLL_INTERVAL = 0.2

# How many times to retry connecting to the scanner
CONNECT_MAX_RETRIES = 5
CONNECT_RETRY_DELAY = 3.0  # seconds between retries

# How many consecutive errors before attempting reconnect
MAX_CONSECUTIVE_ERRORS = 10


# ---------------------------------------------------
# LOGGING SETUP
# ---------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)

logger = logging.getLogger("barcode_system")


# ---------------------------------------------------
# EXCEL MANAGEMENT CLASS
# ---------------------------------------------------

class ExcelManager:
    """
    Handles Excel file creation and writing.

    Column layout per row:
      Col 1 : Wing_ID
      Col 2 : Timestamp
      Col 3,4  : Barcode_1, Inspection_1
      Col 5,6  : Barcode_2, Inspection_2
      ...

    Wing rows start at row 2 (row 1 = headers).
    Wing IDs are 1-based, so wing N → row N+1.
    """

    WING_ID_COLUMN = 1
    TIMESTAMP_COLUMN = 2
    FIRST_BARCODE_COLUMN = 3  # Barcode_1 starts here

    def __init__(self, filename: str):
        self.filename = filename
        self._ensure_file_exists()

    # ---------------------------------------------------

    def _ensure_file_exists(self):
        """Create Excel file with headers if it doesn't exist."""
        if os.path.exists(self.filename):
            logger.info(f"Excel file found: {self.filename}")
            return

        logger.info(f"Creating new Excel file: {self.filename}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Barcodes"

        headers = ["Wing_ID", "Timestamp"]
        for i in range(1, MAX_BARCODES_PER_WING + 1):
            headers.append(f"Barcode_{i}")
            headers.append(f"Inspection_{i}")

        ws.append(headers)
        wb.save(self.filename)

    # ---------------------------------------------------

    def _barcode_column(self, slot: int) -> int:
        """
        Return the Excel column index for a given 1-based slot number.

        Layout: Wing_ID | Timestamp | Barcode_1 | Inspection_1 | Barcode_2 | ...
        Columns:    1   |     2     |     3     |      4       |     5     | ...

        For slot N:
          barcode_col    = 3 + (N - 1) * 2
          inspection_col = 3 + (N - 1) * 2 + 1
        """
        if not (1 <= slot <= MAX_BARCODES_PER_WING):
            raise ValueError(f"Slot {slot} out of range (1-{MAX_BARCODES_PER_WING})")
        return self.FIRST_BARCODE_COLUMN + (slot - 1) * 2

    # ---------------------------------------------------

    def write_barcode(self, wing_id: int, slot: int, barcode: str, inspection: str):
        """
        Write barcode and inspection result for a given wing + slot.

        Wing IDs are 1-based. Row = wing_id + 1 (offset by header row).
        """
        wb = load_workbook(self.filename)
        ws = wb.active

        # Row 1 = headers; wing 1 → row 2, wing N → row N+1
        row = wing_id + 1

        barcode_col = self._barcode_column(slot)
        inspection_col = barcode_col + 1

        ws.cell(row=row, column=self.WING_ID_COLUMN).value = wing_id
        ws.cell(row=row, column=self.TIMESTAMP_COLUMN).value = datetime.now()
        ws.cell(row=row, column=barcode_col).value = barcode
        ws.cell(row=row, column=inspection_col).value = inspection

        wb.save(self.filename)

        logger.info(
            f"Excel updated | Wing {wing_id} | Slot {slot} | "
            f"Barcode {barcode} | {inspection}"
        )


# ---------------------------------------------------
# MODBUS SCANNER INTERFACE
# ---------------------------------------------------

class ScannerInterface:
    """
    Handles Modbus TCP communication with the Cognex scanner.

    Reconnects automatically if the connection is lost.
    """

    def __init__(self, ip: str, port: int):
        self.ip = ip
        self.port = port
        self.client: ModbusTcpClient | None = None

    # ---------------------------------------------------

    def connect(self, max_retries: int = CONNECT_MAX_RETRIES):
        """
        Connect to the scanner with retry logic.

        Raises RuntimeError if all retries are exhausted.
        """
        for attempt in range(1, max_retries + 1):
            logger.info(
                f"Connecting to scanner at {self.ip}:{self.port} "
                f"(attempt {attempt}/{max_retries})..."
            )
            try:
                self.client = ModbusTcpClient(self.ip, port=self.port)
                if self.client.connect():
                    logger.info("Scanner connected successfully.")
                    return
            except Exception as e:
                logger.warning(f"Connection attempt {attempt} failed: {e}")

            if attempt < max_retries:
                logger.info(f"Retrying in {CONNECT_RETRY_DELAY}s...")
                time.sleep(CONNECT_RETRY_DELAY)

        raise RuntimeError(
            f"Could not connect to scanner at {self.ip}:{self.port} "
            f"after {max_retries} attempts."
        )

    # ---------------------------------------------------

    def disconnect(self):
        """Cleanly close the Modbus connection."""
        if self.client:
            self.client.close()
            self.client = None
            logger.info("Scanner disconnected.")

    # ---------------------------------------------------

    def reconnect(self):
        """Disconnect and reconnect."""
        logger.warning("Attempting to reconnect to scanner...")
        self.disconnect()
        self.connect()

    # ---------------------------------------------------

    def _check_connected(self):
        """Raise if client is not initialised."""
        if self.client is None:
            raise RuntimeError("Scanner is not connected.")

    # ---------------------------------------------------

    def scan_available(self) -> bool:
        """Return True if the scan flag register is set to 1."""
        self._check_connected()

        result = self.client.read_holding_registers(SCAN_FLAG_REGISTER, 1)

        if result.isError():
            raise ModbusException("Error reading scan flag register.")

        return result.registers[0] == 1

    # ---------------------------------------------------

    def read_barcode(self) -> str | None:
        """
        Read barcode ASCII characters from holding registers.

        Returns the barcode string, or None if the read failed.
        """
        self._check_connected()

        result = self.client.read_holding_registers(
            BARCODE_START_REGISTER,
            BARCODE_REGISTER_LENGTH
        )

        if result.isError():
            raise ModbusException("Error reading barcode registers.")

        barcode = ""
        for reg in result.registers:
            if reg == 0:
                break
            barcode += chr(reg)

        return barcode.strip() or None

    # ---------------------------------------------------

    def reset_scan_flag(self):
        """Clear the scanner ready flag."""
        self._check_connected()
        self.client.write_register(SCANNER_RESET_REGISTER, 0)


# ---------------------------------------------------
# INSPECTION SYSTEM
# ---------------------------------------------------

class InspectionEngine:
    """
    Determines inspection result for a scanned barcode.

    Rules:
      - None or empty → FAIL
      - Shorter than MIN_LENGTH characters → FAIL
      - Starts with a prefix in INVALID_PREFIXES → FAIL
      - Otherwise → PASS
    """

    MIN_LENGTH = 3
    INVALID_PREFIXES = ("X", "Z")

    def inspect(self, barcode: str | None) -> str:
        if not barcode:
            return "FAIL"

        if len(barcode) < self.MIN_LENGTH:
            return "FAIL"

        if barcode.startswith(self.INVALID_PREFIXES):
            return "FAIL"

        return "PASS"


# ---------------------------------------------------
# DUPLICATE DETECTOR
# ---------------------------------------------------

class DuplicateDetector:
    """
    Prevents duplicate barcode entries within a session.

    Note: the seen-set is in-memory only and resets on process restart.
    If persistence across restarts is required, back this with a file or DB.
    """

    def __init__(self):
        self._seen: set[str] = set()

    # ---------------------------------------------------

    def is_duplicate(self, barcode: str) -> bool:
        """Return True if barcode was already seen; register it if not."""
        if barcode in self._seen:
            return True
        self._seen.add(barcode)
        return False

    # ---------------------------------------------------

    def reset(self):
        """Clear all seen barcodes (e.g. at the start of a new batch)."""
        self._seen.clear()
        logger.info("Duplicate detector reset.")


# ---------------------------------------------------
# SLOT MANAGER
# ---------------------------------------------------

class SlotManager:
    """
    Tracks the current slot (1-16) and wing number (1-based).

    Advances to the next slot on each call to next_slot().
    When slot 16 is passed, wraps to slot 1 and increments the wing.
    """

    def __init__(self):
        self.current_slot = 1
        self.current_wing = 1

    # ---------------------------------------------------

    def next_slot(self):
        """Advance to the next slot; start a new wing if the current one is full."""
        self.current_slot += 1
        if self.current_slot > MAX_BARCODES_PER_WING:
            self.current_slot = 1
            self.current_wing += 1
            logger.info(f"Wing {self.current_wing - 1} complete. Starting wing {self.current_wing}.")

    # ---------------------------------------------------

    @property
    def slot(self) -> int:
        return self.current_slot

    @property
    def wing(self) -> int:
        return self.current_wing


# ---------------------------------------------------
# MAIN CONTROLLER
# ---------------------------------------------------

class BarcodeSystem:
    """
    Main system controller.

    Orchestrates scanning, inspection, duplicate detection, slot
    management, and Excel logging. Handles connection errors by
    attempting to reconnect automatically.
    """

    def __init__(self):
        self.excel = ExcelManager(EXCEL_FILE)
        self.scanner = ScannerInterface(SCANNER_IP, MODBUS_PORT)
        self.inspector = InspectionEngine()
        self.duplicates = DuplicateDetector()
        self.slots = SlotManager()
        self._consecutive_errors = 0

    # ---------------------------------------------------

    def start(self):
        """Connect to the scanner and begin the main polling loop."""
        logger.info("Starting barcode logging system.")
        self.scanner.connect()

        while True:
            try:
                self._poll_once()
                self._consecutive_errors = 0  # reset on success

            except ModbusException as e:
                self._consecutive_errors += 1
                logger.error(f"Modbus error ({self._consecutive_errors}): {e}")

                if self._consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
                    logger.warning("Too many consecutive Modbus errors — reconnecting.")
                    self._safe_reconnect()
                    self._consecutive_errors = 0
                else:
                    time.sleep(2)

            except Exception as e:
                logger.exception(f"Unexpected error: {e}")
                time.sleep(2)

    # ---------------------------------------------------

    def _poll_once(self):
        """Single poll iteration: check for scan, read, inspect, log."""
        if not self.scanner.scan_available():
            time.sleep(SCAN_POLL_INTERVAL)
            return

        barcode = self.scanner.read_barcode()

        if not barcode:
            logger.warning("Scan flag was set but barcode is empty — skipping.")
            self.scanner.reset_scan_flag()
            return

        if self.duplicates.is_duplicate(barcode):
            logger.warning(f"Duplicate barcode ignored: {barcode}")
            self.scanner.reset_scan_flag()
            return

        inspection = self.inspector.inspect(barcode)

        self.excel.write_barcode(
            wing_id=self.slots.wing,
            slot=self.slots.slot,
            barcode=barcode,
            inspection=inspection,
        )

        # Reset scanner AFTER successful write so we don't lose the scan
        # if the Excel write fails (exception would skip this line).
        self.scanner.reset_scan_flag()

        self.slots.next_slot()

    # ---------------------------------------------------

    def _safe_reconnect(self):
        """Attempt reconnect, sleeping between tries. Does not raise."""
        try:
            self.scanner.reconnect()
        except RuntimeError as e:
            logger.error(f"Reconnect failed: {e}. Will retry on next error cycle.")


# ---------------------------------------------------
# PROGRAM ENTRY
# ---------------------------------------------------

def main():
    system = BarcodeSystem()
    system.start()


if __name__ == "__main__":
    main()